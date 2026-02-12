import openpyxl
from openpyxl.styles import PatternFill


class XLUtils:

    @staticmethod
    def load_sheet(file, sheet_name):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheet_name]
        return workbook, sheet


    @staticmethod
    def get_row_count(file, sheet_name):
        workbook, sheet = XLUtils.load_sheet(file, sheet_name)
        count = sheet.max_row
        workbook.close()
        return count


    @staticmethod
    def get_column_count(file, sheet_name):
        workbook, sheet = XLUtils.load_sheet(file, sheet_name)
        count = sheet.max_column
        workbook.close()
        return count


    @staticmethod
    def read_data(file, sheet_name, row, col):
        workbook, sheet = XLUtils.load_sheet(file, sheet_name)
        data = sheet.cell(row=row, column=col).value
        workbook.close()
        return data


    @staticmethod
    def write_data(file, sheet_name, row, col, data):
        workbook, sheet = XLUtils.load_sheet(file, sheet_name)
        sheet.cell(row=row, column=col).value = data
        workbook.save(file)
        workbook.close()


    @staticmethod
    def fill_color(file, sheet_name, row, col, color):
        workbook, sheet = XLUtils.load_sheet(file, sheet_name)

        fill = PatternFill(start_color=color,
                           end_color=color,
                           fill_type="solid")

        sheet.cell(row=row, column=col).fill = fill

        workbook.save(file)
        workbook.close()
